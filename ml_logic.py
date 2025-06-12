import pandas as pd
import numpy as np
import logging
from typing import Any, Dict, List, Tuple, Optional, Union
from pathlib import Path
import io
import re
from datetime import datetime
from decimal import Decimal, getcontext, InvalidOperation, DivisionByZero, DivisionUndefined
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import openpyxl
from difflib import SequenceMatcher
from fuzzywuzzy import fuzz
import os

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Set decimal precision and rounding mode
getcontext().prec = 6
getcontext().rounding = 'ROUND_HALF_UP'

# Confidence thresholds
CONFIDENCE_THRESHOLDS = {
    'high': 0.98,    # Green
    'medium': 0.95,  # Yellow
    'low': 0.0       # Red
}

# Confidence colors
CONFIDENCE_COLORS = {
    'high': 'C6EFCE',    # Green
    'medium': 'FFEB9C',  # Yellow
    'low': 'FFC7CE'      # Red
}

class ConfidenceScore:
    """Confidence scoring engine for field validation."""
    
    @staticmethod
    def get_confidence_level(score: float) -> str:
        """Get confidence level based on score."""
        if score >= CONFIDENCE_THRESHOLDS['high']:
            return 'high'
        elif score >= CONFIDENCE_THRESHOLDS['medium']:
            return 'medium'
        return 'low'
    
    @staticmethod
    def get_color(score: float) -> str:
        """Get color based on confidence score."""
        level = ConfidenceScore.get_confidence_level(score)
        return CONFIDENCE_COLORS[level]
    
    @staticmethod
    def score_pack_size(pack_str: str) -> Tuple[float, Optional[str]]:
        """
        Score pack size format confidence.
        Returns (score, error_message)
        """
        if not isinstance(pack_str, str) or not pack_str.strip():
            return 0.0, "Empty or invalid pack size"
            
        try:
            # Perfect match patterns
            perfect_patterns = [
                r"^\d+\s*[xX*]\s*[\d\.]+\s*(ml|l|fl\.oz|gal|cl)$",  # e.g., "12 x 750ml"
                r"^[\d\.]+\s*(ml|l|fl\.oz|gal|cl)$"                 # e.g., "19.5L"
            ]
            
            # Partial match patterns
            partial_patterns = [
                r"^\d+[xX*][\d\.]+(ml|l|fl\.oz|gal|cl)$",  # e.g., "12x750ml"
                r"^[\d\.]+(ml|l|fl\.oz|gal|cl)$"           # e.g., "19.5l"
            ]
            
            # Check perfect matches
            for pattern in perfect_patterns:
                if re.match(pattern, pack_str.strip()):
                    return 1.0, None
                    
            # Check partial matches
            for pattern in partial_patterns:
                if re.match(pattern, pack_str.strip()):
                    return 0.8, "Non-standard format"
                    
            # Try to extract any numbers and units
            if re.search(r"[\d\.]+", pack_str) and re.search(r"(ml|l|fl\.oz|gal|cl)", pack_str, re.I):
                return 0.5, "Unrecognized format"
                
            return 0.0, "Invalid format"
            
        except Exception as e:
            logger.error(f"Error scoring pack size: {str(e)}")
            return 0.0, str(e)
    
    @staticmethod
    def score_supplier(vendor: str, known_suppliers: List[str]) -> Tuple[float, Optional[str]]:
        """
        Score supplier mapping confidence.
        Returns (score, error_message)
        """
        if not isinstance(vendor, str) or not vendor.strip():
            return 0.0, "Empty or invalid vendor"
            
        try:
            vendor = vendor.strip()
            
            # Exact match
            if vendor in known_suppliers:
                return 1.0, None
                
            # Fuzzy match
            best_score = 0
            best_match = None
            
            for supplier in known_suppliers:
                score = fuzz.ratio(vendor.lower(), supplier.lower())
                if score > best_score:
                    best_score = score
                    best_match = supplier
                    
            if best_score >= 80:
                return 0.9, f"Fuzzy match: {best_match}"
            elif best_score >= 60:
                return 0.7, f"Partial match: {best_match}"
                
            return 0.3, "No close match found"
            
        except Exception as e:
            logger.error(f"Error scoring supplier: {str(e)}")
            return 0.0, str(e)
    
    @staticmethod
    def score_case_size(case_size: float) -> Tuple[float, Optional[str]]:
        """
        Score case size confidence.
        Returns (score, error_message)
        """
        try:
            if not isinstance(case_size, (int, float)):
                return 0.0, "Invalid type"
                
            if case_size <= 0:
                return 0.0, "Case size must be positive"
                
            if case_size > 100:  # Unusually large case size
                return 0.7, "Unusually large case size"
                
            return 1.0, None
            
        except Exception as e:
            logger.error(f"Error scoring case size: {str(e)}")
            return 0.0, str(e)
    
    @staticmethod
    def score_cu_price(cu_price: float, amount: float, quantity: float) -> Tuple[float, Optional[str]]:
        """
        Score CU PRICE confidence.
        Returns (score, error_message)
        """
        try:
            if not isinstance(cu_price, (int, float)):
                return 0.0, "Invalid type"
                
            if cu_price <= 0:
                return 0.0, "CU PRICE must be positive"
                
            # Check if CU PRICE equals case price
            if abs(cu_price - amount) < 0.01:
                return 0.5, "CU PRICE equals case price"
                
            # Check if CU PRICE is reasonable
            if cu_price > amount:
                return 0.3, "CU PRICE exceeds case price"
                
            return 1.0, None
            
        except Exception as e:
            logger.error(f"Error scoring CU PRICE: {str(e)}")
            return 0.0, str(e)
    
    @staticmethod
    def score_unit_of_measure(unit: str) -> Tuple[float, Optional[str]]:
        """
        Score unit of measure confidence.
        Returns (score, error_message)
        """
        if not isinstance(unit, str):
            return 0.0, "Invalid type"
            
        unit = unit.strip().upper()
        
        if unit == "L":
            return 1.0, None
        elif unit in ["ML", "LITER", "LITERS"]:
            return 0.8, "Non-standard unit format"
        else:
            return 0.0, "Invalid unit"

class FieldValidator:
    """Field-level validation and confidence scoring."""
    
    def __init__(self, known_suppliers: List[str]):
        self.known_suppliers = known_suppliers
        self.confidence_scores = {}
        
    def validate_row(self, row: Dict[str, Any]) -> Dict[str, Tuple[float, Optional[str]]]:
        """Validate all fields in a row and return confidence scores."""
        scores = {}
        
        # Validate pack size
        pack_size = row.get('pack_size', '')
        scores['pack_size'] = ConfidenceScore.score_pack_size(pack_size)
        
        # Validate supplier
        vendor = row.get('vendor', '')
        scores['supplier'] = ConfidenceScore.score_supplier(vendor, self.known_suppliers)
        
        # Validate case size
        case_size = row.get('case_size')
        scores['case_size'] = ConfidenceScore.score_case_size(case_size)
        
        # Validate CU PRICE
        cu_price = row.get('cu_price')
        amount = row.get('amount')
        quantity = row.get('quantity')
        scores['cu_price'] = ConfidenceScore.score_cu_price(cu_price, amount, quantity)
        
        # Validate unit of measure
        unit = row.get('unit_of_measure', '')
        scores['unit_of_measure'] = ConfidenceScore.score_unit_of_measure(unit)
        
        return scores

class ConfidenceDashboard:
    """Create and manage confidence dashboard."""
    
    def __init__(self, writer: pd.ExcelWriter):
        self.writer = writer
        self.stats = {
            'total_rows': 0,
            'fields': {}
        }
        
    def update_stats(self, field: str, score: float):
        """Update statistics for a field."""
        if field not in self.stats['fields']:
            self.stats['fields'][field] = {
                'total': 0,
                'valid': 0,
                'invalid': 0,
                'errors': []
            }
            
        self.stats['fields'][field]['total'] += 1
        
        if score >= CONFIDENCE_THRESHOLDS['high']:
            self.stats['fields'][field]['valid'] += 1
        else:
            self.stats['fields'][field]['invalid'] += 1
        
    def create_dashboard(self):
        """Create confidence dashboard sheet."""
        ws = self.writer.book.create_sheet("Dashboard")
        
        # Add title
        ws['A1'] = "Confidence Dashboard"
        ws['A1'].font = Font(bold=True, size=14)
        
        # Add headers
        headers = ['Field', 'Total Values', 'Valid Values', 'Invalid Values', 
                  'Accuracy %', 'Status', 'Common Issues']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            
        # Add data rows
        row = 3
        for field, stats in self.stats['fields'].items():
            total = stats['total']
            if total == 0:
                continue
                
            valid_pct = stats['valid'] / total * 100
            invalid_pct = stats['invalid'] / total * 100
            
            # Determine status
            if valid_pct >= 98:
                status = 'High'
                color = CONFIDENCE_COLORS['high']
            elif valid_pct >= 95:
                status = 'Medium'
                color = CONFIDENCE_COLORS['medium']
            else:
                status = 'Low'
                color = CONFIDENCE_COLORS['low']
            
            # Add row data
            ws.cell(row=row, column=1).value = field
            ws.cell(row=row, column=2).value = total
            ws.cell(row=row, column=3).value = stats['valid']
            ws.cell(row=row, column=4).value = stats['invalid']
            ws.cell(row=row, column=5).value = f"{valid_pct:.1f}%"
            ws.cell(row=row, column=6).value = status
            
            # Add color fill for status
            ws.cell(row=row, column=6).fill = PatternFill(
                start_color=color,
                end_color=color,
                fill_type='solid'
            )
            
            # Add common issues
            if stats['errors']:
                common_issues = pd.Series(stats['errors']).value_counts().head(3)
                issues_text = "\n".join(f"{issue} ({count})" for issue, count in common_issues.items())
                ws.cell(row=row, column=7).value = issues_text
                ws.cell(row=row, column=7).alignment = Alignment(wrap_text=True)
            
            row += 1
            
        # Add summary row
        ws.cell(row=row, column=1).value = "Overall"
        ws.cell(row=row, column=1).font = Font(bold=True)
        
        total_valid = sum(stats['valid'] for stats in self.stats['fields'].values())
        total_values = sum(stats['total'] for stats in self.stats['fields'].values())
        overall_accuracy = (total_valid / total_values * 100) if total_values > 0 else 0
        
        ws.cell(row=row, column=2).value = total_values
        ws.cell(row=row, column=3).value = total_valid
        ws.cell(row=row, column=4).value = total_values - total_valid
        ws.cell(row=row, column=5).value = f"{overall_accuracy:.1f}%"
        
        # Set overall status
        if overall_accuracy >= 98:
            status = 'High'
            color = CONFIDENCE_COLORS['high']
        elif overall_accuracy >= 95:
            status = 'Medium'
            color = CONFIDENCE_COLORS['medium']
        else:
            status = 'Low'
            color = CONFIDENCE_COLORS['low']
            
        ws.cell(row=row, column=6).value = status
        ws.cell(row=row, column=6).fill = PatternFill(
            start_color=color,
            end_color=color,
            fill_type='solid'
        )
        
        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].auto_size = True
            
        # Set row height for wrapped text
        for row in range(3, row + 1):
            ws.row_dimensions[row].height = 40

class PostProcessor:
    """Post-process ML output against reference data."""
    
    def __init__(self, reference_file: str):
        self.reference_file = reference_file
        self.reference_df = None
        self.validation_results = {
            'total_rows': 0,
            'matched_rows': 0,
            'columns': {}
        }
        
    def load_reference_data(self):
        """Load and normalize reference data."""
        try:
            # Check if file exists
            if not os.path.exists(self.reference_file):
                raise FileNotFoundError(f"Reference file not found: {self.reference_file}")
                
            self.reference_df = pd.read_csv(self.reference_file)
            
            # Normalize column names
            self.reference_df.columns = self.reference_df.columns.str.strip().str.upper()
            
            # Normalize string columns
            for col in ['ITEM', 'STORE', 'PACK SIZE', 'Supplier', 'VENDOR']:
                if col in self.reference_df.columns:
                    self.reference_df[col] = self.reference_df[col].astype(str).str.strip().str.upper()
                    
            # Convert numeric columns
            for col in ['Case Size', 'Container Size', 'CU PRICE', 'Total Cases', 'QUANTITY', 'AMOUNT']:
                if col in self.reference_df.columns:
                    self.reference_df[col] = pd.to_numeric(self.reference_df[col], errors='coerce')
                    
            logger.info(f"Loaded reference data: {len(self.reference_df)} rows")
            
        except FileNotFoundError as e:
            logger.error(f"Reference file error: {str(e)}")
            raise
        except Exception as e:
            logger.error(f"Error loading reference data: {str(e)}")
            raise
            
    def normalize_value(self, value: Any, column: str) -> Any:
        """Normalize value for comparison."""
        if pd.isna(value) or value is None:
            return None
            
        if column in ['ITEM', 'STORE', 'PACK SIZE', 'Supplier', 'VENDOR', 'CATEGORY', 'Unit of Measure']:
            return str(value).strip().upper()
            
        if column in ['Case Size', 'Container Size', 'CU PRICE', 'Total Cases', 'QUANTITY', 'AMOUNT']:
            try:
                return float(value)
            except (ValueError, TypeError):
                return None
                
        return value
        
    def values_match(self, val1: Any, val2: Any, column: str) -> Tuple[bool, str]:
        """Check if two values match based on column type."""
        if pd.isna(val1) or pd.isna(val2):
            return (pd.isna(val1) and pd.isna(val2), "Missing value")
            
        if column in ['Case Size', 'Container Size', 'CU PRICE', 'Total Cases', 'QUANTITY', 'AMOUNT']:
            try:
                val1 = float(val1)
                val2 = float(val2)
                # Allow 0.01% difference for numeric values
                if abs(val1 - val2) / max(abs(val1), abs(val2)) < 0.0001:
                    return True, ""
                return False, f"Value mismatch: {val1} vs {val2}"
            except (ValueError, TypeError):
                return False, "Invalid numeric value"
                
        val1_str = str(val1).strip().upper()
        val2_str = str(val2).strip().upper()
        if val1_str == val2_str:
            return True, ""
        return False, f"Value mismatch: {val1_str} vs {val2_str}"
        
    def validate_output(self, output_df: pd.DataFrame) -> pd.DataFrame:
        """Validate output against reference data."""
        if self.reference_df is None:
            self.load_reference_data()
            
        # Initialize validation results
        self.validation_results = {
            'total_rows': len(output_df),
            'matched_rows': 0,
            'columns': {col: {'total': 0, 'matched': 0, 'errors': []} for col in output_df.columns}
        }
        
        # Create a copy of output DataFrame for highlighting
        validated_df = output_df.copy()
        
        # Add validation columns
        validated_df['_match_key'] = validated_df.apply(
            lambda row: f"{row['ITEM']}|{row['STORE']}|{row['PACK SIZE']}", axis=1
        )
        
        # Process each row
        for idx, row in validated_df.iterrows():
            match_key = row['_match_key']
            ref_rows = self.reference_df[
                (self.reference_df['ITEM'] == row['ITEM']) &
                (self.reference_df['STORE'] == row['STORE']) &
                (self.reference_df['PACK SIZE'] == row['PACK SIZE'])
            ]
            
            if len(ref_rows) > 0:
                ref_row = ref_rows.iloc[0]
                self.validation_results['matched_rows'] += 1
                
                # Compare each column
                for col in validated_df.columns:
                    if col.startswith('_'):
                        continue
                        
                    if col in ref_row.index:
                        self.validation_results['columns'][col]['total'] += 1
                        
                        output_val = self.normalize_value(row[col], col)
                        ref_val = self.normalize_value(ref_row[col], col)
                        
                        is_match, error_msg = self.values_match(output_val, ref_val, col)
                        
                        if is_match:
                            self.validation_results['columns'][col]['matched'] += 1
                        else:
                            # Mark cell for highlighting
                            validated_df.at[idx, f'_{col}_highlight'] = True
                            validated_df.at[idx, f'_{col}_error'] = error_msg
                            self.validation_results['columns'][col]['errors'].append(
                                f"Row {idx + 1}: {error_msg}"
                            )
                            
        # Remove temporary columns
        validated_df = validated_df.drop(['_match_key'] + [col for col in validated_df.columns if col.startswith('_')], axis=1)
        
        return validated_df
        
    def create_validation_dashboard(self, writer: pd.ExcelWriter):
        """Create validation dashboard sheet."""
        ws = writer.book.create_sheet("Validation Dashboard")
        
        # Add title
        ws['A1'] = "Validation Results vs Reference File"
        ws['A1'].font = Font(bold=True, size=14)
        
        # Add headers
        headers = ['Column', 'Total Values', 'Matched Values', 'Accuracy %', 'Status', 'Common Issues']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            
        # Add data rows
        row = 3
        for col, stats in self.validation_results['columns'].items():
            total = stats['total']
            if total == 0:
                continue
                
            matched = stats['matched']
            accuracy = (matched / total * 100) if total > 0 else 0
            
            # Determine status
            if accuracy == 100:
                status = 'Perfect'
                color = 'C6EFCE'  # Green
            elif accuracy >= 95:
                status = 'Good'
                color = 'FFEB9C'  # Yellow
            else:
                status = 'Needs Review'
                color = 'FFC7CE'  # Red
            
            # Add row data
            ws.cell(row=row, column=1).value = col
            ws.cell(row=row, column=2).value = total
            ws.cell(row=row, column=3).value = matched
            ws.cell(row=row, column=4).value = f"{accuracy:.1f}%"
            ws.cell(row=row, column=5).value = status
            
            # Add color fill for status
            ws.cell(row=row, column=5).fill = PatternFill(
                start_color=color,
                end_color=color,
                fill_type='solid'
            )
            
            # Add common issues
            if stats['errors']:
                common_issues = pd.Series(stats['errors']).value_counts().head(3)
                issues_text = "\n".join(f"{issue} ({count})" for issue, count in common_issues.items())
                ws.cell(row=row, column=6).value = issues_text
                ws.cell(row=row, column=6).alignment = Alignment(wrap_text=True)
            
            row += 1
            
        # Add summary row
        ws.cell(row=row, column=1).value = "Overall"
        ws.cell(row=row, column=1).font = Font(bold=True)
        
        total_matched = sum(stats['matched'] for stats in self.validation_results['columns'].values())
        total_values = sum(stats['total'] for stats in self.validation_results['columns'].values())
        overall_accuracy = (total_matched / total_values * 100) if total_values > 0 else 0
        
        ws.cell(row=row, column=2).value = total_values
        ws.cell(row=row, column=3).value = total_matched
        ws.cell(row=row, column=4).value = f"{overall_accuracy:.1f}%"
        
        # Set overall status
        if overall_accuracy == 100:
            status = 'Perfect'
            color = 'C6EFCE'  # Green
        elif overall_accuracy >= 95:
            status = 'Good'
            color = 'FFEB9C'  # Yellow
        else:
            status = 'Needs Review'
            color = 'FFC7CE'  # Red
            
        ws.cell(row=row, column=5).value = status
        ws.cell(row=row, column=5).fill = PatternFill(
            start_color=color,
            end_color=color,
            fill_type='solid'
        )
        
        # Add legend
        row += 2
        ws.cell(row=row, column=1).value = "Legend"
        ws.cell(row=row, column=1).font = Font(bold=True)
        
        legend = [
            ("Perfect (100%)", 'C6EFCE'),
            ("Good (95-99%)", 'FFEB9C'),
            ("Needs Review (<95%)", 'FFC7CE')
        ]
        
        for i, (text, color) in enumerate(legend, 1):
            row += 1
            ws.cell(row=row, column=1).value = text
            ws.cell(row=row, column=1).fill = PatternFill(
                start_color=color,
                end_color=color,
                fill_type='solid'
            )
        
        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].auto_size = True

# Define standard column order
STANDARD_COLUMNS = [
    'ITEM', 'STORE', 'VENDOR', 'CATEGORY', 'PACK SIZE', 
    'CASE SIZE', 'UNIT OF MEASURE', 'QUANTITY', 'CU PRICE', 'AMOUNT'
]

# Expected columns for 60 Bev format
BEV_COLS = [
    'STORE', 'SKU', 'ITEM', 'PACK SIZE', 'CATEGORY', 'VENDOR',
    'INVOICE NO', 'RECEIVED DATE', 'INVOICE DATE', 'QUANTITY', 'AMOUNT', 'CU PRICE'
]

# Target columns for 60 Vines format with data types
VINES_COLS = {
    'ITEM': str,
    'STORE': str,
    'VENDOR': str,
    'CATEGORY': str,
    'PACK SIZE': str,
    'CASE SIZE': float,
    'UNIT OF MEASURE': str,
    'QUANTITY': float,
    'AMOUNT': float,
    'CU PRICE': float
}

# Enhanced store mapping
STORE_MAP = {
    # 60 Vines locations
    "sixty vines charlotte": "Sixty Vines Charlotte",
    "sixty vines dallas": "Sixty Vines Dallas",
    "sixty vines houston": "Sixty Vines Houston",
    "sixty vines austin": "Sixty Vines Austin",
    "sixty vines plano": "Sixty Vines Plano",
    "sixty vines irvine": "Sixty Vines Irvine",
    "sixty vines san antonio": "Sixty Vines San Antonio",
    "sixty vines tampa": "Sixty Vines Tampa",
    "sixty vines orlando": "Sixty Vines Orlando",
    "sixty vines miami": "Sixty Vines Miami",
    
    # Common variations
    "60 vines": "Sixty Vines",
    "60vines": "Sixty Vines",
    "sixty vines": "Sixty Vines",
    "sixtyvines": "Sixty Vines",
    
    # Store locations
    "bevmo 1": "BevMo #1",
    "ralphs la": "Ralphs - LA",
    "costco la": "Costco - Los Angeles",
    "wholefoods santa monica": "Whole Foods - Santa Monica",
    "target dtla": "Target - DTLA"
}

# Enhanced supplier mapping with common patterns
SUPPLIER_MAP = {
    # Wine distributors
    "winebow": "Winebow",
    "southern": "Southern Glazer's",
    "youngs": "Young's Market",
    "reyes": "Reyes Beverage",
    "breakthru": "Breakthru Beverage",
    "republic": "Republic National",
    "advintage": "Advintage Distributing",
    "3 keys": "3 KEYS DISTRIBUTING",
    "empire": "Empire Distributors",
    "allied": "Allied Beverage",
    "prestige": "Prestige Beverage",
    
    # Wine producers
    "wine warehouse": "Wine Warehouse",
    "wine merchants": "Wine Merchants",
    "wine & spirits": "Wine & Spirits",
    "wine & beer": "Wine & Beer",
    "wine & more": "Wine & More",
    
    # Common variations
    "southern glazers": "Southern Glazer's",
    "southern glazer": "Southern Glazer's",
    "young's": "Young's Market",
    "youngs market": "Young's Market",
    "reyes beverage": "Reyes Beverage",
    "breakthru beverage": "Breakthru Beverage",
    "republic national": "Republic National",
    "advintage distributing": "Advintage Distributing",
    "3 keys distributing": "3 KEYS DISTRIBUTING",
    "empire distributors": "Empire Distributors",
    "allied beverage": "Allied Beverage",
    "prestige beverage": "Prestige Beverage"
}

# Enhanced category mapping with more keywords
CATEGORY_MAP = {
    # Spirits
    "tequila": "Spirits", "vodka": "Spirits", "gin": "Spirits", "rum": "Spirits",
    "whiskey": "Spirits", "bourbon": "Spirits", "scotch": "Spirits", "mezcal": "Spirits",
    "cognac": "Spirits", "brandy": "Spirits", "liqueur": "Spirits", "absinthe": "Spirits",
    "single malt": "Spirits", "blended": "Spirits", "aged": "Spirits", "reserve": "Spirits",
    "premium": "Spirits", "craft": "Spirits",
    
    # Wine
    "cabernet": "Wine", "merlot": "Wine", "pinot": "Wine", "sauvignon": "Wine",
    "chardonnay": "Wine", "malbec": "Wine", "syrah": "Wine", "shiraz": "Wine",
    "zinfandel": "Wine", "prosecco": "Wine", "champagne": "Wine", "sparkling": "Wine",
    "rose": "Wine", "red blend": "Wine", "white blend": "Wine", "vintage": "Wine",
    "reserve": "Wine", "grand cru": "Wine", "premier cru": "Wine", "brut": "Wine",
    "extra dry": "Wine", "semi-sweet": "Wine", "sweet": "Wine",
    
    # Beer
    "lager": "Beer", "ipa": "Beer", "pilsner": "Beer", "ale": "Beer",
    "stout": "Beer", "porter": "Beer", "sour": "Beer", "wheat": "Beer",
    "hefeweizen": "Beer", "kolsch": "Beer", "saison": "Beer", "cider": "Beer",
    "craft": "Beer", "microbrew": "Beer", "import": "Beer", "domestic": "Beer",
    "light": "Beer", "dark": "Beer", "amber": "Beer",
    
    # Other
    "water": "Other", "juice": "Other", "soda": "Other", "kombucha": "Other",
    "tonic": "Other", "soda water": "Other", "seltzer": "Other", "mixer": "Other",
    "non-alcoholic": "Other", "energy": "Other", "tea": "Other", "coffee": "Other"
}

# Unit conversions to milliliters
UNIT_CONVERSIONS = {
    'ml': 1,
    'l': 1000,
    'oz': 29.5735,
    'gal': 3785.41
}

# Define exact output columns from 60 Vines format
VINES_OUTPUT_COLUMNS = [
    'ITEM', 'STORE', 'Supplier', 'PACK SIZE', 'CATEGORY', 'Case Size',
    'Container Size', 'VENDOR', 'Total Cases', 'QUANTITY', 'AMOUNT',
    'Unit of Measure', 'CU PRICE'
]

# Define column data types and validation rules
COLUMN_RULES = {
    'ITEM': {'type': str, 'required': True},
    'STORE': {'type': str, 'required': True},
    'Supplier': {'type': str, 'required': True},
    'PACK SIZE': {'type': str, 'required': True},
    'CATEGORY': {'type': str, 'required': True},
    'Case Size': {'type': float, 'required': True, 'min': 0},
    'Container Size': {'type': str, 'required': True, 'pattern': r'^\d+\.?\d*L$'},
    'VENDOR': {'type': str, 'required': True},
    'Total Cases': {'type': float, 'required': True, 'min': 0},
    'QUANTITY': {'type': float, 'required': True, 'min': 0},
    'AMOUNT': {'type': float, 'required': True, 'min': 0},
    'Unit of Measure': {'type': str, 'required': True, 'values': ['L']},
    'CU PRICE': {'type': float, 'required': True, 'min': 0, 'max': 200}
}

# Price thresholds for validation
PRICE_THRESHOLDS = {
    'min': 1.0,    # Minimum expected price
    'max': 200.0,  # Maximum expected price
    'outlier': 500.0  # Absolute maximum before flagging
}

def validate_column_value(value: Any, column: str) -> Tuple[bool, Optional[str]]:
    """
    Validate a column value against its rules.
    Returns (is_valid, error_message)
    """
    rules = COLUMN_RULES.get(column, {})
    
    # Check required
    if rules.get('required', False) and (value is None or value == ''):
        return False, f"{column} is required"
        
    # Skip validation for empty values if not required
    if not rules.get('required', False) and (value is None or value == ''):
        return True, None
        
    # Check type
    expected_type = rules.get('type', str)
    if not isinstance(value, expected_type):
        try:
            value = expected_type(value)
        except (ValueError, TypeError):
            return False, f"{column} must be {expected_type.__name__}"
            
    # Check numeric ranges
    if expected_type in (int, float):
        if rules.get('min') is not None and value < rules['min']:
            return False, f"{column} must be >= {rules['min']}"
        if rules.get('max') is not None and value > rules['max']:
            return False, f"{column} must be <= {rules['max']}"
            
    # Check string patterns
    if expected_type == str and rules.get('pattern'):
        if not re.match(rules['pattern'], str(value)):
            return False, f"{column} format is invalid"
            
    # Check allowed values
    if expected_type == str and rules.get('values'):
        if str(value) not in rules['values']:
            return False, f"{column} must be one of {rules['values']}"
            
    return True, None

def format_container_size(size_ml: float) -> str:
    """Format container size in ml to display format."""
    if size_ml is None:
        return ""
        
    # For 19.5L and 20L kegs, show in ml
    if size_ml >= 19000:  # 19.5L or 20L kegs
        return f"{int(size_ml)}"
        
    # For other sizes, show in ml
    return f"{int(size_ml)}"

def calculate_cu_price(amount: float, quantity: float, pack_size_str: str) -> float:
    """Calculate CU PRICE based on pack size and quantity."""
    if amount is None or quantity is None or quantity == 0:
        return 0
        
    # For all sizes, calculate price per unit
    return round(amount / quantity, 2)

def get_confidence_score(validation_results: Dict[str, Tuple[bool, Optional[str]]]) -> float:
    """Calculate overall confidence score from validation results."""
    if not validation_results:
        return 0.0
        
    valid_count = sum(1 for is_valid, _ in validation_results.values() if is_valid)
    return valid_count / len(validation_results)

def standardize_column_names(df: pd.DataFrame) -> pd.DataFrame:
    """
    Standardize column names by stripping whitespace and converting to uppercase.
    More lenient approach that handles variations in column names.
    """
    try:
        # Create a mapping of normalized column names
        column_mapping = {}
        for col in df.columns:
            # Normalize the column name
            normalized = str(col).strip().upper()
            # Handle common variations
            normalized = normalized.replace(' ', '_')
            normalized = normalized.replace('-', '_')
            normalized = normalized.replace('.', '_')
            # Map the original column to the normalized name
            column_mapping[col] = normalized
            
        # Apply the mapping
        df.columns = [column_mapping[col] for col in df.columns]
        
        # Log the cleaned column names
        logging.info(f"Cleaned column names: {df.columns.tolist()}")
        
        return df
    except Exception as e:
        logging.error(f"Error standardizing column names: {str(e)}")
        raise

def normalize_string(s: Any) -> str:
    """Normalize string values for consistent matching with fallback."""
    try:
        if pd.isna(s) or s is None:
            return ""
        return str(s).strip().lower()
    except Exception as e:
        logging.warning(f"Error normalizing string value: {str(e)}")
        return ""

def clean_input_data(df: pd.DataFrame) -> pd.DataFrame:
    """
    Clean and standardize input DataFrame with more lenient validation.
    """
    try:
        # Clean string columns
        for col in df.select_dtypes(include='object').columns:
            try:
                df[col] = df[col].astype(str).str.strip().str.lower()
            except Exception as e:
                logging.warning(f"Error cleaning column {col}: {str(e)}")
                continue
        
        # Clean numeric columns with fallback
        numeric_cols = ['QUANTITY', 'AMOUNT', 'CU_PRICE', 'TOTAL_CASES']
        for col in numeric_cols:
            if col in df.columns:
                try:
                    df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
                except Exception as e:
                    logging.warning(f"Error converting {col} to numeric: {str(e)}")
                    df[col] = 0
        
        # Add normalized columns for matching with fallback
        try:
            df["ITEM_NORM"] = df["ITEM"].apply(normalize_string)
        except Exception as e:
            logging.warning(f"Error normalizing ITEM column: {str(e)}")
            df["ITEM_NORM"] = ""
            
        try:
            df["STORE_NORM"] = df["STORE"].apply(normalize_string)
        except Exception as e:
            logging.warning(f"Error normalizing STORE column: {str(e)}")
            df["STORE_NORM"] = ""
            
        try:
            df["VENDOR_NORM"] = df["VENDOR"].apply(normalize_string)
        except Exception as e:
            logging.warning(f"Error normalizing VENDOR column: {str(e)}")
            df["VENDOR_NORM"] = ""
        
        # Remove any completely empty rows
        df = df.dropna(how='all')
        
        logging.info(f"Cleaned input data: {len(df)} rows")
        return df
        
    except Exception as e:
        logging.error(f"Error cleaning input data: {str(e)}")
        raise

def normalize_store(store: str) -> str:
    """Normalize store name with enhanced mapping."""
    try:
        if pd.isna(store) or not store:
            return ""
            
        store = str(store).lower().strip()
        
        # Try exact match first
        if store in STORE_MAP:
            return STORE_MAP[store]
            
        # Try partial match
        for key, value in STORE_MAP.items():
            if key in store:
                logger.info(f"Mapped store '{store}' to '{value}'")
                return value
                
        # Fallback to title case
        return store.title()
        
    except Exception as e:
        logger.error(f"Error normalizing store '{store}': {str(e)}")
        return store.title()

def map_supplier(vendor: str) -> str:
    """
    Map vendor to supplier using enhanced matching.
    
    Examples:
        "Winebow" -> "Winebow"
        "Southern Glazers" -> "Southern Glazer's"
        "Young's Market" -> "Young's Market"
    """
    try:
        if pd.isna(vendor) or not vendor:
            return "Other"
            
        vendor = str(vendor).lower().strip()
        
        # Try exact match first
        if vendor in SUPPLIER_MAP:
            return SUPPLIER_MAP[vendor]
            
        # Try partial match
        for key, value in SUPPLIER_MAP.items():
            if key in vendor:
                logger.info(f"Mapped vendor '{vendor}' to supplier '{value}'")
                return value
                
        # Fallback to title case
        return vendor.title()
        
    except Exception as e:
        logger.error(f"Error mapping vendor '{vendor}': {str(e)}")
        return "Other"

def categorize_item(item: str) -> str:
    """Categorize item with enhanced keyword matching."""
    try:
        if pd.isna(item) or not item:
            return "Other"
            
        item = str(item).lower()
        matches = []
        
        # Check for matches
        for keyword, category in CATEGORY_MAP.items():
            if keyword in item:
                matches.append((category, len(keyword)))
        
        if matches:
            # Return category with longest matching keyword
            best_match = max(matches, key=lambda x: x[1])
            logger.info(f"Categorized '{item}' as '{best_match[0]}' (match: {best_match[1]} chars)")
            return best_match[0]
        
        logger.warning(f"No category match found for item: {item}")
        return "Other"
        
    except Exception as e:
        logger.error(f"Error categorizing item '{item}': {str(e)}")
        return "Other"

def safe_decimal(value: Any) -> Decimal:
    """
    Safely convert a value to Decimal.
    Returns Decimal("0") if conversion fails.
    """
    try:
        if pd.isna(value) or value is None:
            return Decimal("0")
        if isinstance(value, str) and not value.strip():
            return Decimal("0")
        return Decimal(str(value))
    except (InvalidOperation, TypeError, ValueError) as e:
        logger.warning(f"Failed to convert {value} to Decimal: {e}")
        return Decimal("0")

def parse_pack_size(pack_str: str) -> Tuple[Optional[int], Optional[float]]:
    """Parse pack size string to extract case size and container size in ml."""
    if not pack_str or not isinstance(pack_str, str):
        return None, None
        
    pack_str = pack_str.lower().strip()
    
    # Special handling for 19.5L and 20L kegs
    if '19.5l' in pack_str or '19.5 l' in pack_str:
        return 2.17, 19500  # 19.5L kegs use 2.17 case size
    if '20l' in pack_str or '20 l' in pack_str:
        return 2.22, 20000  # 20L kegs use 2.22 case size
        
    # Handle standard bottle formats
    patterns = [
        (r'(\d+)\s*x\s*(\d+(?:\.\d+)?)\s*(ml|l|oz)', 2),  # 6 x 750ml, 6 x 1.5l
        (r'(\d+(?:\.\d+)?)\s*(ml|l|oz)', 1),  # 750ml, 1.5l
        (r'(\d+)\s*/\s*(\d+(?:\.\d+)?)\s*(ml|l|oz)', 2),  # 6/750ml
        (r'(\d+)\s*-\s*(\d+(?:\.\d+)?)\s*(ml|l|oz)', 2),  # 6-750ml
    ]
    
    for pattern, group_idx in patterns:
        match = re.search(pattern, pack_str, re.IGNORECASE)
        if match:
            try:
                if group_idx == 1:
                    # Single unit
                    size = float(match.group(1))
                    unit = match.group(2).lower()
                    # Convert to ml
                    if unit == 'l':
                        size = size * 1000
                    elif unit == 'oz':
                        size = size * 29.5735
                    return 1, size
                else:
                    # Case format
                    case_size = int(match.group(1))
                    size = float(match.group(2))
                    unit = match.group(3).lower()
                    # Convert to ml
                    if unit == 'l':
                        size = size * 1000
                    elif unit == 'oz':
                        size = size * 29.5735
                    return case_size, size
            except (ValueError, IndexError):
                continue
    
    # If no pattern matches, try to extract numbers and units
    numbers = re.findall(r'\d+(?:\.\d+)?', pack_str)
    units = re.findall(r'ml|l|oz', pack_str, re.IGNORECASE)
    
    if numbers and units:
        try:
            if len(numbers) == 1:
                size = float(numbers[0])
                unit = units[0].lower()
                if unit == 'l':
                    size = size * 1000
                elif unit == 'oz':
                    size = size * 29.5735
                return 1, size
            elif len(numbers) >= 2:
                case_size = int(numbers[0])
                size = float(numbers[1])
                unit = units[0].lower()
                if unit == 'l':
                    size = size * 1000
                elif unit == 'oz':
                    size = size * 29.5735
                return case_size, size
        except (ValueError, IndexError):
            pass
    
    return None, None

def calculate_case_size(container_size_ml: float, case_size: int) -> Optional[float]:
    """Calculate case size in ml."""
    if container_size_ml is None or case_size is None:
        return None
    return container_size_ml * case_size

def calculate_total_cases(quantity: float, case_size: int, pack_size_str: str) -> float:
    """Calculate total cases based on pack size."""
    if quantity is None or case_size is None or case_size == 0:
        return 0
        
    # Special handling for 19.5L and 20L kegs
    if '19.5l' in pack_size_str.lower() or '19.5 l' in pack_size_str.lower():
        return round(quantity * 2.17, 2)  # 19.5L kegs use 2.17 multiplier
    if '20l' in pack_size_str.lower() or '20 l' in pack_size_str.lower():
        return round(quantity * 2.22, 2)  # 20L kegs use 2.22 multiplier
        
    # For standard bottles, use case_size
    return round(quantity / case_size, 2)

def calculate_amount(cu_price: Decimal, quantity: Decimal) -> Decimal:
    """Calculate amount with full precision using Decimal."""
    try:
        cu_price = safe_decimal(cu_price)
        quantity = safe_decimal(quantity)
        
        amount = cu_price * quantity
        logger.info(f"Calculated AMOUNT: {amount} (CU PRICE: {cu_price}, QUANTITY: {quantity})")
        return amount
    except Exception as e:
        logger.error(f"Error calculating AMOUNT: {str(e)}")
        return safe_decimal('0')

def safe_case_calc(quantity: float, case_size: float) -> float:
    """Safely calculate total cases."""
    try:
        return round(quantity / case_size, 2) if case_size and case_size > 0 else 0
    except:
        return 0

def process_bev_to_vines(input_df: pd.DataFrame, writer: pd.ExcelWriter) -> str:
    """Process 60 Bev data and transform to 60 Vines format with exact column matching."""
    try:
        # Initialize error tracking and confidence dashboard
        error_rows = []
        skipped_rows = 0
        dashboard = ConfidenceDashboard(writer)
        validator = FieldValidator(list(SUPPLIER_MAP.values()))
        
        # Normalize column headers
        input_df.columns = input_df.columns.str.strip().str.lower().str.replace(" ", "_")
        
        # Initialize output DataFrame with exact columns
        output_df = pd.DataFrame(columns=VINES_OUTPUT_COLUMNS)
        
        # Store validation results for each cell
        cell_validations = {}  # {(row_idx, col_name): (is_valid, error_msg)}
        
        # Process each row
        for idx, row in input_df.iterrows():
            try:
                # Get values with fallback
                item = str(row.get('item', '')).strip()
                store = str(row.get('store', 'N/A')).strip()
                vendor = str(row.get('vendor', '')).strip()
                supplier = map_supplier(vendor)  # Use enhanced supplier mapping
                pack_size = str(row.get('pack_size', '')).strip()
                category = str(row.get('category', 'Unknown')).strip()
                quantity = float(row.get('quantity', 0))
                amount = round(float(row.get('amount', row.get('total', row.get('extended_price', 0)))), 2)
                
                # Log row being processed
                logger.info(f"Processing row {idx}: {store} | {vendor}")
                
                # Parse pack size
                case_size, container_size_ml = parse_pack_size(pack_size)
                if case_size is None or container_size_ml is None:
                    case_size = 1
                    container_size_ml = 750  # Default to 750ml if parsing fails
                    logger.warning(f"Using default values for row {idx}: {pack_size}")
                    cell_validations[(idx, 'PACK SIZE')] = (False, "Using default values")
                
                # Calculate derived values
                case_size_ml = calculate_case_size(container_size_ml, case_size)
                total_cases = calculate_total_cases(quantity, case_size, pack_size)
                cu_price = calculate_cu_price(amount, quantity, pack_size)
                
                # Create output row with exact column names
                output_row = {
                    'ITEM': item,
                    'STORE': store,
                    'Supplier': supplier,
                    'PACK SIZE': pack_size,
                    'CATEGORY': category,
                    'CASE SIZE': case_size,
                    'CONTAINER SIZE': format_container_size(container_size_ml),
                    'VENDOR': vendor,
                    'TOTAL CASES': total_cases,
                    'QUANTITY': quantity,
                    'AMOUNT': amount,
                    'CU PRICE': cu_price
                }
                
                # Validate all fields in the output row
                for field, value in output_row.items():
                    is_valid, error_msg = validate_column_value(value, field)
                    if not is_valid:
                        cell_validations[(idx, field)] = (False, error_msg)
                        logger.warning(f"Row {idx}, {field}: {error_msg}")
                        # Update confidence dashboard with actual validation results
                        dashboard.update_stats(field, 0.0)  # Mark as invalid
                    else:
                        dashboard.update_stats(field, 1.0)  # Mark as valid
                
                # Add row to output DataFrame
                output_df = pd.concat([output_df, pd.DataFrame([output_row])], ignore_index=True)
                
            except Exception as e:
                error_rows.append({
                    'row': idx,
                    'error': str(e),
                    'store': str(row.get('store', '')),
                    'vendor': str(row.get('vendor', ''))
                })
                logger.error(f"Error processing row {idx}: {str(e)}")
                skipped_rows += 1
        
        # Create error DataFrame
        error_df = pd.DataFrame(error_rows)
        
        # Save to Excel
        try:
            # Save main data with confidence highlighting
            output_df.to_excel(writer, sheet_name="60 Vines Output", index=False)
            
            # Apply confidence highlighting
            ws = writer.sheets["60 Vines Output"]
            
            # Create yellow highlight style
            yellow_fill = PatternFill(
                start_color='FFFF00',  # Yellow
                end_color='FFFF00',
                fill_type='solid'
            )
            
            # Add header row for validation messages
            header_row = len(output_df) + 2
            ws.cell(row=header_row, column=1).value = "Validation Messages"
            ws.cell(row=header_row, column=1).font = Font(bold=True)
            
            # Highlight cells and add validation messages
            for idx, row in output_df.iterrows():
                for col, field in enumerate(output_df.columns, 1):
                    # Get validation result for this cell
                    validation_result = cell_validations.get((idx, field), (True, None))
                    is_valid, error_msg = validation_result
                    
                    if not is_valid:
                        # Highlight cell in yellow
                        cell = ws.cell(row=idx + 2, column=col)  # +2 for header row
                        cell.fill = yellow_fill
                        
                        # Add validation message in the message row
                        msg_cell = ws.cell(row=header_row, column=col)
                        if msg_cell.value is None:
                            msg_cell.value = f"Row {idx + 1}: {error_msg}"
                        else:
                            msg_cell.value += f"\nRow {idx + 1}: {error_msg}"
                        msg_cell.alignment = Alignment(wrap_text=True)
            
            # Auto-adjust column widths
            for col in range(1, len(output_df.columns) + 1):
                ws.column_dimensions[get_column_letter(col)].auto_size = True
            
            # Create confidence dashboard
            dashboard.create_dashboard()
            
            # Save debug log if there are errors
            if not error_df.empty:
                error_df.to_excel(writer, sheet_name='Debug_Log', index=False)
                
                # Auto-adjust column widths for debug log
                ws = writer.sheets['Debug_Log']
                for idx, col in enumerate(error_df.columns):
                    max_length = max(
                        error_df[col].astype(str).apply(len).max(),
                        len(str(col))
                    )
                    ws.column_dimensions[get_column_letter(idx + 1)].width = max_length + 2
                
        except Exception as e:
            logger.error(f"Error saving Excel file: {str(e)}")
            # Try to save at least the main data
            output_df.to_excel(writer, sheet_name="60 Vines Output", index=False)
        
        # Create summary message
        summary = [
            f"Total rows processed: {len(input_df)}",
            f"Valid rows: {len(output_df)}",
            f"Skipped rows: {skipped_rows}",
            f"Rows with errors: {len(error_df)}",
            "\n⚠️ Cells highlighted in yellow need manual review. See validation messages at the bottom of the sheet."
        ]
        
        if not error_df.empty:
            summary.append("\n⚠️ Some rows could not be fully processed. See 'Debug_Log' tab in output file.")
        
        return "\n".join(summary)
        
    except Exception as e:
        logger.error(f"Error in process_bev_to_vines: {str(e)}")
        # Return minimal output with error message
        pd.DataFrame({'Error': [str(e)]}).to_excel(writer, sheet_name="Error", index=False)
        return f"Error processing file: {str(e)}"

def clean_column_names(df: pd.DataFrame) -> pd.DataFrame:
    """Clean column names to be more lenient with validation."""
    try:
        # Create a mapping of normalized column names
        column_mapping = {}
        for col in df.columns:
            # Normalize the column name
            normalized = str(col).strip().lower()
            # Handle common variations
            normalized = normalized.replace(' ', '_')
            normalized = normalized.replace('-', '_')
            normalized = normalized.replace('.', '_')
            # Map the original column to the normalized name
            column_mapping[col] = normalized
            
        # Apply the mapping
        df.columns = [column_mapping[col] for col in df.columns]
        
        # Log the cleaned column names
        logger.info(f"Cleaned column names: {df.columns.tolist()}")
        
        return df
    except Exception as e:
        logger.error(f"Error standardizing column names: {str(e)}")
        return df  # Return original DataFrame if cleaning fails

def read_input_file(file_data: bytes, file_extension: str) -> pd.DataFrame:
    """Read input file based on extension."""
    try:
        if file_extension.lower() == '.csv':
            return pd.read_csv(io.BytesIO(file_data))
        else:
            return pd.read_excel(io.BytesIO(file_data))
    except Exception as e:
        logger.error(f"Error reading input file: {str(e)}")
        raise

def read_reference_file(file_data: bytes, file_extension: str) -> pd.DataFrame:
    """Read reference file based on extension."""
    try:
        if file_extension.lower() == '.csv':
            return pd.read_csv(io.BytesIO(file_data))
        else:
            return pd.read_excel(io.BytesIO(file_data))
    except Exception as e:
        logger.error(f"Error reading reference file: {str(e)}")
        raise

def is_bev_format(df: pd.DataFrame) -> bool:
    """Check if the DataFrame matches 60 Bev format."""
    try:
        # Check if all required columns are present
        required_cols = set(BEV_COLS)
        actual_cols = set(df.columns)
        
        if not required_cols.issubset(actual_cols):
            missing_cols = required_cols - actual_cols
            logger.warning(f"Missing required columns: {missing_cols}")
            return False
        
        return True
    except Exception as e:
        logger.error(f"Error checking format: {str(e)}")
        return False

def clean_text_values(df: pd.DataFrame) -> pd.DataFrame:
    """Clean text values to title case and strip whitespace."""
    for col in ['ITEM', 'STORE', 'VENDOR']:
        if col in df.columns:
            df[col] = df[col].astype(str).str.strip().str.title()
    return df 