"""
BBB Purchase Log Normalizer
Processes purchase data with supplier matching and confidence scoring.
"""

import pandas as pd
import numpy as np
import os
import io
from typing import Dict, List, Tuple, Optional
from fuzzywuzzy import fuzz
import logging
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import re
from io import BytesIO
from fuzzywuzzy import process
import traceback

# Import utility functions
try:
    from utils.file_utils import (
        read_excel_file,
        read_csv_file,
        save_excel_file,
        generate_output_filename,
        parse_date_column,
        normalize_column_names
    )
except ImportError:
    # Fallback for when running as module
    import sys
    sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    from utils.file_utils import (
        read_excel_file,
        read_csv_file,
        save_excel_file,
        generate_output_filename,
        parse_date_column,
        normalize_column_names
    )

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class BBBNormalizer:
    """BBB Purchase Log Normalizer with Supplier Matching"""
    
    def __init__(self, supplier_reference_file: str):
        """
        Initialize with supplier reference file
        Args:
            supplier_reference_file: Path to Excel file with supplier item list
        """
        self.supplier_reference_file = supplier_reference_file
        self.supplier_items = []
        self.supplier_mapping = {}
        self.load_supplier_reference()
        
    def load_supplier_reference(self):
        """Load and parse supplier reference file"""
        try:
            # Check if file exists
            if not os.path.exists(self.supplier_reference_file):
                logger.error(f"Supplier reference file not found: {self.supplier_reference_file}")
                logger.error(f"Current working directory: {os.getcwd()}")
                logger.error(f"Available files in test_files: {os.listdir('test_files') if os.path.exists('test_files') else 'test_files directory not found'}")
                raise FileNotFoundError(f"Supplier reference file not found: {self.supplier_reference_file}")
            
            # Read supplier reference file using utility function
            supplier_df = read_excel_file(self.supplier_reference_file)
            
            # Normalize column names using utility function
            supplier_df = normalize_column_names(supplier_df)
            logger.info(f"Supplier reference columns after normalization: {list(supplier_df.columns)}")
            
            # Find the item/product description column
            item_col = None
            for col in supplier_df.columns:
                if 'item' in col or 'product' in col:
                    item_col = col
                    break
            if not item_col:
                logger.error("No item/product description column found in supplier reference file!")
                self.supplier_items = []
                return
            
            # Filter out summary/total rows
            before_rows = len(supplier_df)
            supplier_df = supplier_df[(supplier_df[item_col].notna()) & (~supplier_df['supplier'].astype(str).str.lower().str.contains('total'))]
            after_rows = len(supplier_df)
            logger.info(f"Filtered out {before_rows - after_rows} summary/total rows from supplier reference file.")
            
            # Store the full reference dataframe for supplier mapping
            self.supplier_reference_df = supplier_df
            
            # Extract supplier items from the item column
            self.supplier_items = supplier_df[item_col].dropna().unique().tolist()
            self.supplier_item_col = item_col
            self.supplier_supplier_col = 'supplier'
            logger.info(f"Loaded {len(self.supplier_items)} reference items for matching from column: {item_col}")
        except Exception as e:
            logger.error(f"Error loading supplier reference: {str(e)}")
            self.supplier_items = []
    
    def detect_schema_and_remap(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Intelligent schema detection and column mapping to standard format
        """
        # Normalize column names using utility function
        df = normalize_column_names(df)
        
        logger.info(f"Original columns: {list(df.columns)}")
        
        # Clean and convert 'total' column to numeric if present
        if 'total' in df.columns:
            df['total'] = pd.to_numeric(df['total'], errors='coerce')
            logger.info("Cleaned and converted 'total' column to numeric")
        
        # Comprehensive column mappings with pattern matching
        column_mappings = {
            # Item mappings (product descriptions)
            'product_description': 'item',
            'product_name': 'item',
            'description': 'item',
            'product': 'item',
            'item_name': 'item',
            'item_description': 'item',
            'name': 'item',
            
            # Store mappings
            'customer_name': 'store',
            'location': 'store',
            'store_name': 'store',
            'customer': 'store',
            'retailer_name': 'store',
            'distributor_name': 'store',
            'client': 'store',
            
            # Vendor/Supplier mappings
            'supplier': 'vendor',
            'supplier_name': 'vendor',
            'distributor': 'vendor',
            'vendor_name': 'vendor',
            'retailer_vendor_id': 'vendor',
            'supplier_id': 'vendor',
            
            # Quantity mappings
            'qty': 'quantity',
            'qty_delivered': 'quantity',
            'units': 'quantity',
            'unit_count': 'quantity',
            'case_count': 'quantity',
            'pack_count': 'quantity',
            
            # Amount/Cost mappings
            'total': 'amount',
            'total_amount': 'amount',
            'invoice_amount': 'amount',
            'cost': 'amount',
            'extended_price': 'amount',
            'invoice_total': 'amount',
            'line_total': 'amount',
            
            # Unit Cost/Price mappings
            'unit_cost': 'cu_price',
            'unit_price': 'cu_price',
            'price': 'cu_price',
            'cost_per_unit': 'cu_price',
            'unit_of_measure': 'unit_measure',
            'uom': 'unit_measure',
            
            # Date mappings
            'invoice_date': 'invoice_date',
            'order_date': 'invoice_date',
            'date': 'invoice_date',
            'process_date': 'invoice_date',
            'received_date': 'received_date',
            'delivery_date': 'received_date',
            'due_date': 'received_date',
            
            # SKU/Product Code mappings
            'product_code': 'sku',
            'item_code': 'sku',
            'code': 'sku',
            'product_number': 'sku',
            'upc_number': 'sku',
            'pack_upc': 'sku',
            
            # Pack Size mappings
            'pack_size': 'pack_size',
            'size': 'pack_size',
            'volume': 'pack_size',
            'capacity': 'pack_size',
            
            # Category mappings
            'category': 'category',
            'product_category': 'category',
            'item_category': 'category',
            'type': 'category',
            'product_type': 'category',
        }
        
        # Apply direct mappings
        for old_col, new_col in column_mappings.items():
            if old_col in df.columns and new_col not in df.columns:
                df[new_col] = df[old_col]
                logger.info(f"Mapped column: {old_col} -> {new_col}")
        
        # Intelligent pattern-based mapping
        self._apply_pattern_mapping(df)
        
        # Extract pack size from item descriptions if not found
        if 'pack_size' not in df.columns and 'item' in df.columns:
            df['pack_size'] = df['item'].apply(self._extract_pack_size_from_description)
            logger.info("Extracted pack size from item descriptions")
        
        # Extract category from item descriptions if not found
        if 'category' not in df.columns and 'item' in df.columns:
            df['category'] = df['item'].apply(self._extract_category_from_description)
            logger.info("Extracted category from item descriptions")
        
        # Extract vendor from other columns if not found
        if 'vendor' not in df.columns:
            vendor_candidates = ['distributor_name', 'retailer_vendor_id', 'supplier_id']
            for candidate in vendor_candidates:
                if candidate in df.columns:
                    df['vendor'] = df[candidate]
                    logger.info(f"Mapped vendor from: {candidate}")
                    break
        
        # Ensure required columns exist with defaults
        required_columns = ['item', 'store', 'vendor', 'pack_size', 'category', 'quantity', 'amount', 'cu_price', 'unit_measure']
        for col in required_columns:
            if col not in df.columns:
                if col == 'pack_size':
                    df[col] = 'Standard'
                elif col == 'category':
                    df[col] = 'General'
                elif col == 'unit_measure':
                    df[col] = 'EA'
                else:
                    df[col] = ''
                logger.info(f"Added default column: {col}")
        
        # Special logic: If 'total' column exists and 'quantity' does not, use it as 'quantity'
        if 'total' in df.columns and 'quantity' not in df.columns:
            df['quantity'] = df['total']
            logger.info("Mapped column: total -> quantity (special logic for Moxies/Second Rodeo style files)")
        
        logger.info(f"Final columns: {list(df.columns)}")
        return df
    
    def _apply_pattern_mapping(self, df: pd.DataFrame):
        """Apply intelligent pattern-based column mapping"""
        for col in df.columns:
            col_lower = col.lower()
            
            # Pack size patterns
            if any(pattern in col_lower for pattern in ['pack', 'size', 'volume', 'ml', 'l', 'oz', 'fl']):
                if 'pack_size' not in df.columns:
                    df['pack_size'] = df[col]
                    logger.info(f"Pattern mapped pack_size from: {col}")
            
            # Category patterns
            elif any(pattern in col_lower for pattern in ['category', 'type', 'class', 'group']):
                if 'category' not in df.columns:
                    df['category'] = df[col]
                    logger.info(f"Pattern mapped category from: {col}")
            
            # Vendor patterns
            elif any(pattern in col_lower for pattern in ['vendor', 'supplier', 'distributor', 'retailer']):
                if 'vendor' not in df.columns:
                    df['vendor'] = df[col]
                    logger.info(f"Pattern mapped vendor from: {col}")
            
            # Unit measure patterns
            elif any(pattern in col_lower for pattern in ['uom', 'unit', 'measure']):
                if 'unit_measure' not in df.columns:
                    df['unit_measure'] = df[col]
                    logger.info(f"Pattern mapped unit_measure from: {col}")
    
    def _extract_pack_size_from_description(self, description: str) -> str:
        """Extract pack size from item description using regex patterns"""
        if pd.isna(description) or not description:
            return 'Standard'
        
        desc = str(description).upper()
        
        # Common pack size patterns
        patterns = [
            r'(\d+(?:\.\d+)?)\s*(ML|L|OZ|FL\.?OZ)',  # 750ml, 1.5L, 12oz, 12fl.oz
            r'(\d+)\s*(PACK|PK|CT|COUNT)',  # 6-pack, 12pk, 24ct
            r'(\d+)\s*(CASE|CS)',  # 6-case, 12cs
            r'(\d+)\s*(BOTTLE|BTL)',  # 6-bottle, 12btl
            r'(\d+)\s*(CAN|BOTTLE)',  # 6-can, 12bottle
            r'(\d+)\s*(BBL|BARREL)',  # 1/2BBL, 1/6BBL
        ]
        
        for pattern in patterns:
            match = re.search(pattern, desc)
            if match:
                size = match.group(1)
                unit = match.group(2)
                return f"{size}{unit}"
        
        # Default pack sizes based on common terms
        if any(term in desc for term in ['KEG', 'BBL', 'BARREL']):
            return '19.5L'
        elif any(term in desc for term in ['BOTTLE', 'BTL']):
            return '750ml'
        elif any(term in desc for term in ['CAN', 'BOTTLE']):
            return '12oz'
        elif any(term in desc for term in ['PACK', 'PK']):
            return '6-pack'
        
        return 'Standard'
    
    def _extract_category_from_description(self, description: str) -> str:
        """Extract category from item description"""
        if pd.isna(description) or not description:
            return 'General'
        
        desc = str(description).upper()
        
        # Category patterns
        if any(term in desc for term in ['WINE', 'PINOT', 'CABERNET', 'MERLOT', 'CHARDONNAY', 'SAUVIGNON']):
            return 'Wine'
        elif any(term in desc for term in ['BEER', 'IPA', 'LAGER', 'ALE', 'STOUT', 'PORTER']):
            return 'Beer'
        elif any(term in term in desc for term in ['VODKA', 'WHISKEY', 'BOURBON', 'TEQUILA', 'GIN', 'RUM', 'SCOTCH']):
            return 'Spirits'
        elif any(term in desc for term in ['CIDER', 'HARD CIDER']):
            return 'Cider'
        elif any(term in desc for term in ['CHAMPAGNE', 'PROSECCO', 'SPARKLING']):
            return 'Sparkling'
        elif any(term in desc for term in ['LIQUEUR', 'AMARO', 'BITTERS']):
            return 'Liqueur'
        elif any(term in desc for term in ['MIXER', 'TONIC', 'SODA']):
            return 'Mixer'
        
        return 'General'
    
    def normalize_data_types(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Normalize data types: dates, currency, etc.
        """
        # Normalize dates using utility function
        date_columns = ['invoice_date', 'received_date']
        for col in date_columns:
            if col in df.columns:
                df = parse_date_column(df, col)
        
        # Normalize currency/amount columns
        amount_columns = ['amount', 'quantity']
        for col in amount_columns:
            if col in df.columns:
                try:
                    # Remove currency symbols and convert to float
                    df[col] = df[col].astype(str).str.replace('$', '').str.replace(',', '')
                    df[col] = pd.to_numeric(df[col], errors='coerce')
                    logger.info(f"Normalized amount column: {col}")
                except Exception as e:
                    logger.warning(f"Could not normalize amount column {col}: {str(e)}")
        
        return df
    
    def _combine_multiple_sheets(self, excel_file: pd.ExcelFile) -> pd.DataFrame:
        """
        Combine multiple sheets from Excel file into a single dataframe
        Handles different column names across sheets (Spirit, Wine, Beer, etc.)
        """
        combined_dfs = []
        
        # Define which sheets to process and their item column mappings
        sheet_configs = {
            'Spirits': 'Spirit',
            'Wine': 'Wine', 
            'Package Beer': 'Beer',
            'Draft Beer': 'Beer',
            'Bar Supplies': 'Item',
            'Non-Alcoholic': 'Item'
        }
        
        for sheet_name in excel_file.sheet_names:
            if sheet_name in sheet_configs:
                try:
                    logger.info(f"Processing sheet: {sheet_name}")
                    sheet_df = pd.read_excel(excel_file, sheet_name)
                    
                    if len(sheet_df) > 0:
                        # Map the item column to standard 'item' name
                        item_col = sheet_configs[sheet_name]
                        if item_col in sheet_df.columns:
                            sheet_df['item'] = sheet_df[item_col]
                            logger.info(f"Added {len(sheet_df)} rows from {sheet_name} sheet")
                            combined_dfs.append(sheet_df)
                        else:
                            logger.warning(f"Item column '{item_col}' not found in {sheet_name} sheet")
                    else:
                        logger.warning(f"Sheet {sheet_name} is empty")
                        
                except Exception as e:
                    logger.warning(f"Error processing sheet {sheet_name}: {e}")
        
        if combined_dfs:
            # Combine all dataframes
            combined_df = pd.concat(combined_dfs, ignore_index=True)
            logger.info(f"Combined {len(combined_df)} total rows from {len(combined_dfs)} sheets")
            return combined_df
        else:
            logger.warning("No valid sheets found, using first sheet")
            return pd.read_excel(excel_file, excel_file.sheet_names[0])
    
    def match_supplier_with_confidence(self, item: str) -> Tuple[str, float, str]:
        """
        Match item to supplier using fuzzy matching
        Returns: (matched_supplier_name, confidence_score, status)
        """
        if not item or pd.isna(item) or str(item).strip() == '':
            return "UNMATCHED", 0.0, "Empty item"
        
        item_str = str(item).strip()
        
        # Exact match first (case insensitive)
        for idx, supplier_item in enumerate(self.supplier_items):
            if supplier_item and supplier_item.lower() == item_str.lower():
                supplier_row = self.supplier_reference_df[self.supplier_reference_df[self.supplier_item_col] == supplier_item]
                if not supplier_row.empty:
                    supplier_name = supplier_row.iloc[0][self.supplier_supplier_col]
                    return supplier_name, 1.0, "Exact match"
        
        # Partial match (check if item contains supplier item or vice versa)
        for supplier_item in self.supplier_items:
            if supplier_item and (item_str.lower() in supplier_item.lower() or supplier_item.lower() in item_str.lower()):
                supplier_row = self.supplier_reference_df[self.supplier_reference_df[self.supplier_item_col] == supplier_item]
                if not supplier_row.empty:
                    supplier_name = supplier_row.iloc[0][self.supplier_supplier_col]
                    return supplier_name, 0.95, "Partial match"
        
        # Fuzzy matching with lower thresholds
        best_match = None
        best_score = 0
        for supplier_item in self.supplier_items:
            if not supplier_item:
                continue
            ratio_score = fuzz.ratio(item_str.lower(), supplier_item.lower())
            partial_score = fuzz.partial_ratio(item_str.lower(), supplier_item.lower())
            token_sort_score = fuzz.token_sort_ratio(item_str.lower(), supplier_item.lower())
            score = max(ratio_score, partial_score, token_sort_score)
            if score > best_score:
                best_score = score
                best_match = supplier_item
        if best_match and best_score >= 80:
            supplier_row = self.supplier_reference_df[self.supplier_reference_df[self.supplier_item_col] == best_match]
            if not supplier_row.empty:
                supplier_name = supplier_row.iloc[0][self.supplier_supplier_col]
                return supplier_name, best_score / 100, "High confidence"
        elif best_match and best_score >= 60:
            supplier_row = self.supplier_reference_df[self.supplier_reference_df[self.supplier_item_col] == best_match]
            if not supplier_row.empty:
                supplier_name = supplier_row.iloc[0][self.supplier_supplier_col]
                return supplier_name, best_score / 100, "Medium confidence"
        elif best_match and best_score >= 40:
            supplier_row = self.supplier_reference_df[self.supplier_reference_df[self.supplier_item_col] == best_match]
            if not supplier_row.empty:
                supplier_name = supplier_row.iloc[0][self.supplier_supplier_col]
                return supplier_name, best_score / 100, "Low confidence"
        return "UNMATCHED", best_score / 100, "No confident match"
    
    def process_input_file(self, input_file: str) -> pd.DataFrame:
        """
        Process input file with schema detection and data normalization
        """
        try:
            # Read input file using utility function with robust parsing
            if input_file.lower().endswith('.csv'):
                try:
                    # Try different encodings and separators
                    df = read_csv_file(input_file)
                except Exception as csv_error:
                    logger.warning(f"Standard CSV parsing failed: {csv_error}, trying alternative methods")
                    try:
                        # Try with different encoding
                        df = pd.read_csv(input_file, encoding='latin-1', on_bad_lines='skip')
                    except:
                        try:
                            # Try with different separator
                            df = pd.read_csv(input_file, sep='\t', on_bad_lines='skip')
                        except:
                            # Last resort: try with engine='python'
                            df = pd.read_csv(input_file, engine='python', on_bad_lines='skip')
            else:
                # Handle multi-sheet Excel files
                try:
                    excel_file = pd.ExcelFile(input_file)
                    if len(excel_file.sheet_names) > 1:
                        logger.info(f"Detected multi-sheet Excel file with {len(excel_file.sheet_names)} sheets: {excel_file.sheet_names}")
                        df = self._combine_multiple_sheets(excel_file)
                    else:
                        df = read_excel_file(input_file)
                except Exception as e:
                    logger.warning(f"Multi-sheet detection failed: {e}, trying single sheet")
                    df = read_excel_file(input_file)
            
            logger.info(f"Loaded input file with {len(df)} rows and columns: {list(df.columns)}")
            
            # Detect schema and remap columns
            df = self.detect_schema_and_remap(df)
            
            # Normalize data types
            df = self.normalize_data_types(df)
            
            # Perform supplier matching for each item
            logger.info("Starting supplier matching...")
            supplier_results = []
            
            for idx, row in df.iterrows():
                item = row.get('item', '')
                matched_supplier, confidence, status = self.match_supplier_with_confidence(item)
                supplier_results.append({
                    'supplier_matched': matched_supplier,
                    'confidence_score': confidence,
                    'match_status': status
                })
                
                # Log first few matches for debugging
                if idx < 5:
                    logger.info(f"Row {idx}: '{item}' -> '{matched_supplier}' (confidence: {confidence:.2f}, status: {status})")
                
                if idx % 1000 == 0:
                    logger.info(f"Processed {idx} rows for supplier matching")
            
            # Add supplier matching results to dataframe
            supplier_df = pd.DataFrame(supplier_results)
            df = pd.concat([df, supplier_df], axis=1)
            
            # Log matching statistics
            unmatched_count = len(supplier_df[supplier_df['supplier_matched'] == 'UNMATCHED'])
            total_count = len(supplier_df)
            logger.info(f"Completed supplier matching for {total_count} items")
            logger.info(f"Unmatched items: {unmatched_count} ({unmatched_count/total_count*100:.1f}%)")
            logger.info(f"Matched items: {total_count - unmatched_count} ({(total_count - unmatched_count)/total_count*100:.1f}%)")
            
            # Debug: Log columns and sample values before fallback
            logger.info(f"[DEBUG] Columns before fallback: {list(df.columns)}")
            logger.info(f"[DEBUG] Sample values before fallback: {df.head(3).to_dict('records')}")

            # Only apply fallback if original input had 'total' but not 'quantity'
            original_cols = set(df.columns)
            if 'total' in original_cols and 'quantity' not in original_cols:
                if 'total' in df.columns:
                    if df['quantity'].isna().all():
                        df['quantity'] = df['total']
                        logger.info("[DEBUG] Fallback: Filled quantity from total column (Moxies/Second Rodeo style)")
                if 'Total Cases' in df.columns:
                    if df['Total Cases'].isna().all():
                        df['Total Cases'] = df['total']
                        logger.info("[DEBUG] Fallback: Filled Total Cases from total column (Moxies/Second Rodeo style)")

            # Debug: Log columns and sample values after fallback
            logger.info(f"[DEBUG] Columns after fallback: {list(df.columns)}")
            logger.info(f"[DEBUG] Sample values after fallback: {df.head(3).to_dict('records')}")

            return df
            
        except Exception as e:
            logger.error(f"Error processing input file: {str(e)}")
            raise
    
    def create_purchase_log_sheet(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Create Purchase Log sheet with exact column format from reference file
        Reference format: ['ITEM      ', 'STORE      ', 'Supplier', 'PACK SIZE      ', 'CATEGORY      ', 'Case Size', 'Container Size', 'VENDOR      ', 'Total Cases', 'QUANTITY      ', 'AMOUNT      ', 'Unit of Measure', 'CU PRICE      ']
        """
        # Start with the processed input data
        purchase_log = df.copy()
        
        # Ensure all required columns exist with exact names from reference
        required_columns = [
            'ITEM      ', 'STORE      ', 'Supplier', 'PACK SIZE      ', 'CATEGORY      ', 
            'Case Size', 'Container Size', 'VENDOR      ', 'Total Cases', 'QUANTITY      ', 
            'AMOUNT      ', 'Unit of Measure', 'CU PRICE      '
        ]
        
        # Map existing columns to required format
        column_mapping = {
            'item': 'ITEM      ',
            'store': 'STORE      ',
            'vendor': 'VENDOR      ',
            'pack_size': 'PACK SIZE      ',
            'category': 'CATEGORY      ',
            'quantity': 'QUANTITY      ',
            'amount': 'AMOUNT      ',
            'cu_price': 'CU PRICE      ',
            'unit_measure': 'Unit of Measure'
        }
        
        # Apply column mappings
        for old_col, new_col in column_mapping.items():
            if old_col in purchase_log.columns:
                purchase_log[new_col] = purchase_log[old_col]
        
        # Add missing columns with proper values
        missing_columns = set(required_columns) - set(purchase_log.columns)
        for col in missing_columns:
            if col == 'Supplier':
                # Use supplier mapping results - access the column directly
                if 'supplier_matched' in purchase_log.columns:
                    purchase_log[col] = purchase_log['supplier_matched']
                else:
                    purchase_log[col] = 'UNMATCHED'
            elif col == 'Case Size':
                # Try to extract case size from pack size or use default
                purchase_log[col] = purchase_log['PACK SIZE      '].apply(self._extract_case_size)
            elif col == 'Container Size':
                # Try to extract container size from pack size or use default
                purchase_log[col] = purchase_log['PACK SIZE      '].apply(self._extract_container_size)
            elif col == 'Total Cases':
                # Use quantity as total cases if available
                if 'QUANTITY      ' in purchase_log.columns:
                    purchase_log[col] = purchase_log['QUANTITY      ']
                else:
                    purchase_log[col] = 1
            elif col == 'Unit of Measure':
                if 'Unit of Measure' in purchase_log.columns:
                    purchase_log[col] = purchase_log['Unit of Measure']
                else:
                    purchase_log[col] = 'EA'
            else:
                purchase_log[col] = ''
        
        # Ensure exact column order
        purchase_log = purchase_log[required_columns]
        
        # Log some sample data to verify the flow
        logger.info(f"Created Purchase Log with {len(purchase_log)} rows and columns: {list(purchase_log.columns)}")
        logger.info(f"Sample Purchase Log data - First 3 rows:")
        logger.info(f"Suppliers: {purchase_log['Supplier'].value_counts().head().to_dict()}")
        logger.info(f"Total Cases sum: {purchase_log['Total Cases'].sum()}")
        
        return purchase_log
    
    def _extract_case_size(self, pack_size: str) -> int:
        """Extract case size from pack size description"""
        if pd.isna(pack_size) or not pack_size:
            return 1
        
        pack_str = str(pack_size).upper()
        
        # Look for case patterns
        case_patterns = [
            r'(\d+)\s*(CASE|CS)',  # 6-case, 12cs
            r'(\d+)\s*(PACK|PK)',  # 6-pack, 12pk
            r'(\d+)\s*(COUNT|CT)',  # 6-count, 12ct
        ]
        
        for pattern in case_patterns:
            match = re.search(pattern, pack_str)
            if match:
                return int(match.group(1))
        
        # Default case size based on common terms
        if any(term in pack_str for term in ['KEG', 'BBL', 'BARREL']):
            return 1  # Kegs are typically single units
        elif any(term in pack_str for term in ['BOTTLE', 'BTL']):
            return 12  # Standard case of bottles
        elif any(term in pack_str for term in ['CAN']):
            return 24  # Standard case of cans
        
        return 1
    
    def _extract_container_size(self, pack_size: str) -> int:
        """Extract container size from pack size description"""
        if pd.isna(pack_size) or not pack_size:
            return 1
        
        pack_str = str(pack_size).upper()
        
        # Look for container patterns
        container_patterns = [
            r'(\d+(?:\.\d+)?)\s*(ML|L|OZ|FL\.?OZ)',  # 750ml, 1.5L, 12oz
        ]
        
        for pattern in container_patterns:
            match = re.search(pattern, pack_str)
            if match:
                size = float(match.group(1))
                unit = match.group(2)
                
                # Convert to standard units (ml)
                if unit in ['L', 'LITER']:
                    return int(size * 1000)
                elif unit in ['OZ', 'FL.OZ', 'FL OZ']:
                    return int(size * 29.5735)  # Convert oz to ml
                else:
                    return int(size)
        
        # Default container sizes
        if any(term in pack_str for term in ['KEG', 'BBL', 'BARREL']):
            return 19500  # 19.5L keg
        elif any(term in pack_str for term in ['BOTTLE', 'BTL']):
            return 750  # Standard 750ml bottle
        elif any(term in pack_str for term in ['CAN']):
            return 355  # Standard 12oz can (355ml)
        
        return 1
    
    def create_item_totals_sheet(self, purchase_log: pd.DataFrame) -> pd.DataFrame:
        """
        Create Item Totals sheet with exact column format from reference file
        Reference format: ['Supplier', 'ITEM      ', 'SUM of Total Cases']
        """
        # Group by Supplier and Item, sum Total Cases
        item_totals = purchase_log.groupby(['Supplier', 'ITEM      '])['Total Cases'].sum().reset_index()
        
        # Rename to match reference format
        item_totals = item_totals.rename(columns={'Total Cases': 'SUM of Total Cases'})
        
        # Ensure exact column order
        required_columns = ['Supplier', 'ITEM      ', 'SUM of Total Cases']
        item_totals = item_totals[required_columns]
        
        # Log verification data
        logger.info(f"Created Item Totals with {len(item_totals)} rows")
        logger.info(f"Item Totals - Total SUM of Total Cases: {item_totals['SUM of Total Cases'].sum()}")
        logger.info(f"Item Totals - Sample data (first 3 rows):")
        if len(item_totals) > 0:
            logger.info(f"Sample: {item_totals.head(3).to_dict('records')}")
        
        return item_totals
    
    def create_supplier_totals_sheet(self, purchase_log: pd.DataFrame) -> pd.DataFrame:
        """
        Create Supplier Totals sheet with exact column format from reference file
        Reference format: ['Supplier', 'SUM of Total Cases', 'Unnamed: 2']
        """
        # Group by Supplier, sum Total Cases
        supplier_totals = purchase_log.groupby('Supplier')['Total Cases'].sum().reset_index()
        
        # Rename to match reference format
        supplier_totals = supplier_totals.rename(columns={'Total Cases': 'SUM of Total Cases'})
        
        # Add the 'Unnamed: 2' column (empty column from reference)
        supplier_totals['Unnamed: 2'] = ''
        
        # Ensure exact column order
        required_columns = ['Supplier', 'SUM of Total Cases', 'Unnamed: 2']
        supplier_totals = supplier_totals[required_columns]
        
        # Log verification data
        logger.info(f"Created Supplier Totals with {len(supplier_totals)} rows")
        logger.info(f"Supplier Totals - Total SUM of Total Cases: {supplier_totals['SUM of Total Cases'].sum()}")
        logger.info(f"Supplier Totals - Sample data (first 3 rows):")
        if len(supplier_totals) > 0:
            logger.info(f"Sample: {supplier_totals.head(3).to_dict('records')}")
        
        return supplier_totals
    
    def save_to_excel(self, purchase_log: pd.DataFrame, item_totals: pd.DataFrame, 
                     supplier_totals: pd.DataFrame, output_path: str):
        """
        Save the three sheets to Excel with proper formatting
        """
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Write each sheet
            purchase_log.to_excel(writer, sheet_name='Purchase Log', index=False)
            item_totals.to_excel(writer, sheet_name='Item Totals', index=False)
            supplier_totals.to_excel(writer, sheet_name='Supplier Totals', index=False)
            
            # Apply formatting to each sheet
            for sheet_name in ['Purchase Log', 'Item Totals', 'Supplier Totals']:
                worksheet = writer.sheets[sheet_name]
                self._apply_sheet_formatting(worksheet)
        
        logger.info(f"Saved output to: {output_path}")

    def normalize(self, file_bytes: bytes, references=None):
        """
        Main normalization method that takes file bytes and returns result dict
        Args:
            file_bytes: Bytes of the uploaded file
            references: Optional reference data (not used in this version)
        Returns:
            dict: Contains filename and summary statistics
        """
        try:
            # Generate output filename
            filename = f"BBB_Normalized_Output_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            output_path = os.path.join("files", filename)
            logger.info(f"Generated output path for BBB: {output_path}")

            # Save uploaded file temporarily with proper extension detection
            file_extension = '.csv'  # Default to CSV
            try:
                # Try to detect if it's an Excel file by checking the first few bytes
                if file_bytes.startswith(b'\x50\x4b\x03\x04') or file_bytes.startswith(b'\x50\x4b\x05\x06'):
                    file_extension = '.xlsx'
                elif file_bytes.startswith(b'\xd0\xcf\x11\xe0'):
                    file_extension = '.xls'
            except:
                pass
            
            temp_input_path = os.path.join("files", f"input_bev_{datetime.now().strftime('%Y%m%d_%H%M%S')}{file_extension}")
            with open(temp_input_path, 'wb') as f:
                f.write(file_bytes)
            
            # Process the input file with robust parsing
            processed_df = self.process_input_file(temp_input_path)
            
            # Create output sheets
            purchase_log = self.create_purchase_log_sheet(processed_df)
            item_totals = self.create_item_totals_sheet(purchase_log)
            supplier_totals = self.create_supplier_totals_sheet(purchase_log)
            
            # Save to Excel
            self.save_to_excel(purchase_log, item_totals, supplier_totals, output_path)
            
            # Clean up temporary file
            if os.path.exists(temp_input_path):
                os.remove(temp_input_path)
            
            # Create summary with NaN handling
            total_cases_sum = purchase_log['Total Cases'].sum()
            avg_cases = purchase_log['Total Cases'].mean()
            
            summary = {
                'total_rows': int(len(purchase_log)),
                'unique_suppliers': int(purchase_log['Supplier'].nunique()),
                'unique_items': int(purchase_log['ITEM      '].nunique()),
                'total_cases': float(total_cases_sum) if pd.notna(total_cases_sum) else 0.0,
                'avg_cases_per_item': float(avg_cases) if pd.notna(avg_cases) else 0.0
            }
            
            logger.info(f"BBB normalization complete. Output saved to: {filename}")
            
            return {
                "success": True,
                "filename": filename,
                "summary": summary
            }
            
        except Exception as e:
            tb = traceback.format_exc()
            logger.error(f"Error in BBBNormalizer.normalize: {e}\n{tb}")
            return {"error": str(e), "traceback": tb}

    def _apply_sheet_formatting(self, worksheet):
        """Apply basic formatting to worksheet"""
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Format header row
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font

def normalize_sku_name(sku_name):
    """Normalize SKU names for better matching"""
    if pd.isna(sku_name):
        return ""
    
    # Convert to string and clean
    sku = str(sku_name).strip().lower()
    
    # Remove common prefixes/suffixes
    sku = re.sub(r'\b(6pk|12pk|24pk|case|bottle|can|pack)\b', '', sku)
    
    # Remove special characters but keep spaces
    sku = re.sub(r'[^\w\s]', ' ', sku)
    
    # Normalize whitespace
    sku = ' '.join(sku.split())
    
    return sku

def normalize_store_name(store_name):
    """Normalize store names for better matching"""
    if pd.isna(store_name):
        return ""
    
    store = str(store_name).strip().lower()
    
    # Remove common store prefixes
    store = re.sub(r'\b(store|location|outlet)\b', '', store)
    
    # Remove special characters but keep spaces
    store = re.sub(r'[^\w\s]', ' ', store)
    
    # Normalize whitespace
    store = ' '.join(store.split())
    
    return store

def extract_upc_last_5(upc_str):
    """Extract last 5 digits from UPC for matching"""
    if pd.isna(upc_str):
        return None
    
    upc = str(upc_str).strip()
    # Extract last 5 digits
    if len(upc) >= 5:
        return upc[-5:]
    return None

def generate_insight(row):
    """Generate insights based on performance metrics"""
    try:
        fulfillment = row.get('fulfillment_pct', 0)
        ros = row.get('ros', 0)
        
        if fulfillment >= 90 and ros >= 50:
            return "High performer - maintain current strategy"
        elif fulfillment >= 80 and ros >= 30:
            return "Good performance - monitor trends"
        elif fulfillment >= 70 and ros >= 20:
            return "Moderate performance - consider optimization"
        elif fulfillment < 70:
            return "Low fulfillment - investigate supply chain"
        elif ros < 20:
            return "Low ROS - review pricing strategy"
        else:
            return "Standard performance - continue monitoring"
    except:
        return "Performance data unavailable"

def classify_performance_enhanced(row):
    """Enhanced classification combining fulfillment and ROS metrics"""
    try:
        fulfillment = row.get('fulfillment_pct', 0)
        ros = row.get('ros', 0)
        
        # Fulfillment tiers
        if fulfillment >= 90:
            fulfillment_tier = "High Fulfillment"
        elif fulfillment >= 70:
            fulfillment_tier = "Medium Fulfillment"
        else:
            fulfillment_tier = "Low Fulfillment"
        
        # ROS tiers
        if ros >= 50:
            ros_tier = "High ROS"
        elif ros >= 20:
            ros_tier = "Medium ROS"
        else:
            ros_tier = "Low ROS"
        
        return f"{fulfillment_tier}, {ros_tier}"
    except:
        return "Unknown"

def fuzzy_match_with_fallback(query, choices, threshold=60, fallback_choices=None):
    """Fuzzy match with fallback to alternative choices"""
    if pd.isna(query) or not query:
        return None, 0
    
    query = str(query).strip().lower()
    
    # Try primary choices
    if choices:
        best_match = process.extractOne(query, choices, scorer=fuzz.token_sort_ratio)
        if best_match and best_match[1] >= threshold:
            return best_match[0], best_match[1]
    
    # Try fallback choices if provided
    if fallback_choices:
        fallback_match = process.extractOne(query, fallback_choices, scorer=fuzz.token_sort_ratio)
        if fallback_match and fallback_match[1] >= threshold:
            return fallback_match[0], fallback_match[1]
    
    return None, 0

def fuzzy_merge(left, right, left_on, right_on, threshold=90, limit=1):
    """Perform fuzzy merge between two dataframes"""
    matches = []
    
    for idx, left_row in left.iterrows():
        left_val = str(left_row[left_on]).lower()
        best_match = process.extractOne(left_val, right[right_on].astype(str).str.lower(), scorer=fuzz.token_sort_ratio)
        
        if best_match and best_match[1] >= threshold:
            right_idx = right[right[right_on].astype(str).str.lower() == best_match[0]].index[0]
            matches.append((idx, right_idx, best_match[1]))
    
    return matches

def robust_header_parse(df: pd.DataFrame, required_cols):
    """Robustly parse headers and find required columns"""
    # Try to find header row
    for i in range(min(10, len(df))):
        row = df.iloc[i]
        if all(str(col).lower() in [str(cell).lower() for cell in row] for col in required_cols):
            # Found header row
            df.columns = df.iloc[i]
            df = df.iloc[i+1:].reset_index(drop=True)
            return df
    
    # If no header found, use first row as header
    df.columns = df.iloc[0]
    df = df.iloc[1:].reset_index(drop=True)
    return df

def detect_header_row(df, expected_cols, max_rows=10):
    """Detect which row contains the header"""
    for i in range(min(max_rows, len(df))):
        row = df.iloc[i]
        row_str = ' '.join(str(cell).lower() for cell in row if pd.notna(cell))
        
        # Check if this row contains expected column names
        matches = sum(1 for col in expected_cols if col.lower() in row_str)
        if matches >= len(expected_cols) * 0.7:  # 70% match threshold
            return i
    
    return 0  # Default to first row

def fuzzy_match(val, choices, threshold=80):
    """Simple fuzzy matching function"""
    if pd.isna(val) or not val:
        return None
    
    val_str = str(val).lower()
    best_match = process.extractOne(val_str, [str(c).lower() for c in choices], scorer=fuzz.token_sort_ratio)
    
    if best_match and best_match[1] >= threshold:
        return choices[best_match[2]]  # Return original case
    return None

def normalize_bbb(bev_bytes, references=None):
    """
    Normalize BBB (Beverage) data with supplier matching.
    Args:
        bev_bytes: Bytes of the uploaded file
        references: Optional reference data (not used in this version)
    Returns:
        dict: Contains filename and summary statistics
    """
    try:
        filename = f"BBB_Normalized_Output_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        output_path = os.path.join("files", filename)
        logger.info(f"Generated output path for BBB: {output_path}")

        # Load the uploaded file
        try:
            df = pd.read_excel(BytesIO(bev_bytes))
        except Exception as e:
            logger.warning(f"Excel load failed: {e}, trying CSV.")
            try:
                df = pd.read_csv(BytesIO(bev_bytes))
            except Exception as e2:
                logger.error(f"Failed to load file as Excel or CSV: {e2}")
                return {"error": "Could not read input file as Excel or CSV."}

        # Normalize column names: lowercase, strip, remove spaces/underscores
        def norm_col(col):
            return str(col).strip().lower().replace(' ', '').replace('_', '')
        df.columns = [norm_col(c) for c in df.columns]
        logger.info(f"Sanitized columns: {list(df.columns)}")

        # Define all possible variants for required columns
        col_map = {
            'item': ['item', 'productdescription', 'description', 'productdesc', 'product', 'itemname'],
            'supplier': ['supplier', 'vendor', 'distributor', 'brand'],
            'quantity': ['quantity', 'qty', 'totalcases', 'cases', 'amount', 'totalqty'],
            'sku': ['sku', 'skuno', 'skunumber', 'sku#', 'itemcode'],
            'store': ['store', 'storename', 'location'],
            'case_size': ['casesize', 'packsize', 'size', 'container', 'containersize'],
        }

        # Find best match for each required column
        found_cols = {}
        for key, variants in col_map.items():
            found = None
            for v in variants:
                if v in df.columns:
                    found = v
                    break
            if found:
                found_cols[key] = found
                logger.info(f"Mapped column: {key} -> {found}")
            else:
                found_cols[key] = None
                logger.warning(f"Missing expected column for '{key}'. Will fill with empty values.")
                df[key] = ''  # Add empty column if missing

        # Always use the mapped or fallback columns
        item_col = found_cols['item'] if found_cols['item'] else 'item'
        supplier_col = found_cols['supplier'] if found_cols['supplier'] else 'supplier'
        quantity_col = found_cols['quantity'] if found_cols['quantity'] else 'quantity'
        sku_col = found_cols['sku'] if found_cols['sku'] else 'sku'
        store_col = found_cols['store'] if found_cols['store'] else 'store'
        case_size_col = found_cols['case_size'] if found_cols['case_size'] else 'case_size'

        # Build Purchase Log with all required columns (fill missing with empty)
        purchase_log_cols = [item_col, store_col, supplier_col, case_size_col, sku_col, quantity_col]
        for col in purchase_log_cols:
            if col not in df.columns:
                df[col] = ''
        purchase_log = df[purchase_log_cols].copy()
        purchase_log.columns = ['ITEM', 'STORE', 'Supplier', 'Case Size', 'SKU', 'QUANTITY']
        logger.info(f"Created Purchase Log with {len(purchase_log)} rows and columns: {list(purchase_log.columns)}")

        # Item Totals
        item_totals = purchase_log.groupby(['Supplier', 'ITEM'])['QUANTITY'].apply(lambda x: pd.to_numeric(x, errors='coerce').sum()).reset_index()
        item_totals = item_totals.rename(columns={'QUANTITY': 'SUM of Total Cases'})
        logger.info(f"Created Item Totals with {len(item_totals)} rows")

        # Supplier Totals
        supplier_totals = purchase_log.groupby('Supplier')['QUANTITY'].apply(lambda x: pd.to_numeric(x, errors='coerce').sum()).reset_index()
        supplier_totals = supplier_totals.rename(columns={'QUANTITY': 'SUM of Total Cases'})
        supplier_totals['Unnamed: 2'] = ''
        logger.info(f"Created Supplier Totals with {len(supplier_totals)} rows")

        # Save to Excel
        with pd.ExcelWriter(output_path) as writer:
            purchase_log.to_excel(writer, sheet_name='Purchase Log', index=False)
            item_totals.to_excel(writer, sheet_name='Item Totals', index=False)
            supplier_totals.to_excel(writer, sheet_name='Supplier Totals', index=False)
        logger.info(f"Saved output to: {output_path}")
        logger.info(f"BBB normalization complete. Output saved to: {filename}")

        # Summary
        summary = {
            'total_rows': int(len(purchase_log)),
            'unique_suppliers': int(purchase_log['Supplier'].nunique()),
            'unique_items': int(purchase_log['ITEM'].nunique()),
            'total_cases': float(pd.to_numeric(purchase_log['QUANTITY'], errors='coerce').sum()),
            'avg_cases_per_item': float(pd.to_numeric(purchase_log['QUANTITY'], errors='coerce').mean() or 0)
        }

        return {
            "success": True,
            "filename": filename,
            "summary": summary
        }

    except Exception as e:
        tb = traceback.format_exc()
        logger.error(f"Error in normalize_bbb: {e}\n{tb}")
        return {"error": str(e), "traceback": tb} 